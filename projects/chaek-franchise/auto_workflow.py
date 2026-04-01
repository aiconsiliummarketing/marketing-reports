#!/usr/bin/env python3
"""
차액가맹금 통합 워크플로우 자동 실행 스크립트
엑셀 파일 경로를 받으면 STEP 1→2→3 + 분석보고서 행 추가까지 전부 자동 실행
"""
import sys, os, json, csv, io, subprocess
from datetime import datetime, date, timedelta
import openpyxl, requests

# ========== 설정 ==========
WEBAPP_URL = "https://script.google.com/macros/s/AKfycbybyf-0X4jd21733SkunLBDA_LU4xm_wlKEeKUr85gVl5wNrjI9UOZm3m7awvVdOfz4Xw/exec"
SHEET_ID = "17_FhUn2v5xp6yXxbGV1KO584yiAYROv8ycMv8Persn8"
GITHUB_REPO = "aiconsiliummarketing/marketing-reports"
REPORT_DIR = "/Users/aiconsilium/.openclaw/workspace/projects/chaek-franchise/reports"
WORKSPACE = "/Users/aiconsilium/.openclaw/workspace"

GIDS = {
    "ga1": 948924274, "ga2": 1648398317, "meta": 1508176585,
    "karrot": 1852345446, "nv_grp": 446628138,
    "nv_kw": 2107644143, "nv_cr": 1880680011, "nv_sq": 27295358,
    "pivot": 2058915904,
}

TARGET_SHEETS = [
    "집단소송플랫폼_애널리틱스_Raw", "차액가맹금테스트페이지_애널리틱스_Raw",
    "메타광고_Raw", "네이버검색광고(광고그룹별)_Raw", "네이버검색광고(키워드별)_Raw",
    "네이버검색광고(소재별)_Raw", "네이버검색광고(검색어별)_Raw",
]

def sf(v):
    try:
        v = str(v).replace(',','').strip()
        if v in ('','-','—','None','#N/A'): return 0.0
        return float(v)
    except: return 0.0

def fmt(n):
    if isinstance(n, float): return f"{n:,.0f}"
    return f"{n:,}" if isinstance(n, int) and n != 0 else str(n)

def chg(cur, prev):
    if prev == 0 and cur == 0: return ""
    if prev == 0: return " ▲" if cur > 0 else ""
    pct = (cur - prev) / prev * 100
    if abs(pct) < 0.05: return ""
    return f" ▲{abs(pct):.1f}%" if pct > 0 else f" ▼{abs(pct):.1f}%"

def classify_channel(sm, camp):
    sm_l, camp_l = str(sm).lower(), str(camp).lower()
    if 'home / popup_coupang' in sm_l or ('kakao' in sm_l and 'coupang' in camp_l): return 'coupang'
    if 'kakao' in sm_l and 'crm_at_franchise' in sm_l: return 'crm'
    if 'kakao' in sm_l and 'classaction_' in camp_l: return 'crm'
    if any(x in sm_l for x in ['meta','facebook','fb','ig / paid','fb_ig','adsmanager.facebook']): return 'meta'
    if any(x in sm_l for x in ['naver / sa','naver / cpc','naver / sa_mo','naver / sa_pc']): return 'naver_sa'
    if any(x in sm_l for x in ['karrot','krt_','krt /']): return 'karrot'
    return 'other'

def dl_sheet(gid):
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={gid}"
    resp = requests.get(url, timeout=60)
    text = resp.content.decode('utf-8')
    reader = csv.reader(io.StringIO(text))
    header = next(reader)
    return header, list(reader)

def determine_mode():
    """KST 기준 요일 판정"""
    from datetime import timezone
    import pytz
    kst = pytz.timezone('Asia/Seoul')
    now = datetime.now(kst)
    weekday = now.weekday()  # 0=월
    if weekday == 0:
        return 'B', now  # 주간
    return 'A', now  # 데일리

# ========== STEP 1: Raw 데이터 입력 ==========
def step1(xlsx_path):
    print("=== STEP 1: Raw 데이터 입력 ===")
    
    COL_MAP = {
        "집단소송플랫폼_애널리틱스_Raw": ["날짜","이벤트 이름","세션 소스/매체","세션 캠페인","총 사용자","이벤트 수","활성 사용자당 이벤트 수","평균 세션 시간","이탈률","참여율","방문수"],
        "차액가맹금테스트페이지_애널리틱스_Raw": ["날짜","이벤트 이름","세션 소스/매체","세션 캠페인","총 사용자","이벤트 수","활성 사용자당 이벤트 수","평균 세션 시간","이탈률","참여율","방문수"],
        "메타광고_Raw": ["캠페인 이름","광고 세트 이름","광고 이름","일","게재 상태","게재 수준","지출 금액 (KRW)","결과 유형","결과","결과당 비용","노출","도달","빈도","CPM(1,000회 노출당 비용)","CTR(전체)","시작","종료","링크 클릭","CPC(링크 클릭당 비용)","결과 비율","CTR(링크 클릭률)","랜딩 페이지 조회","링크 클릭당 랜딩 페이지 조회율","보고 시작","보고 종료"],
        "네이버검색광고(광고그룹별)_Raw": ["캠페인유형","캠페인","일별","광고그룹","노출수","클릭수","클릭률(%)","평균 CPC","총비용"],
        "네이버검색광고(키워드별)_Raw": ["캠페인","광고그룹","키워드","일별","노출수","클릭수","클릭률(%)","평균 CPC","총비용","평균노출순위"],
        "네이버검색광고(소재별)_Raw": ["캠페인","광고그룹","소재","소재 유형","일별","노출수","클릭수","클릭률(%)","평균 CPC","총비용","총 전환수","직접전환수","간접전환수","총 전환율(%)"],
        "네이버검색광고(검색어별)_Raw": ["검색어","검색 유형","일별","캠페인","광고그룹","노출수","클릭수","클릭률(%)","평균 CPC","총비용","평균노출순위"],
    }
    
    payload = {"sheets": []}
    
    if xlsx_path.endswith('.json'):
        # JSON 입력
        with open(xlsx_path, 'r') as f:
            raw = json.load(f)
        for name in TARGET_SHEETS:
            rows_json = raw.get("data", {}).get(name, [])
            if not rows_json: continue
            cols = COL_MAP.get(name, list(rows_json[0].keys()))
            data = [[row.get(c, '') for c in cols] for row in rows_json]
            payload["sheets"].append({"name": name, "data": data})
            print(f"  {name}: {len(data)}행")
    else:
        # 엑셀 입력
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        def fmt_val(val):
            if val is None: return ''
            if isinstance(val, (datetime, date)): return val.strftime('%Y-%m-%d')
            return val
        for name in TARGET_SHEETS:
            if name not in wb.sheetnames: continue
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) <= 1: continue
            data = [r for r in rows[1:] if r and any(c is not None and str(c).strip() != '' for c in r)]
            if not data: continue
            formatted = [[fmt_val(c) for c in row] for row in data]
            payload["sheets"].append({"name": name, "data": formatted})
            print(f"  {name}: {len(formatted)}행")
        wb.close()

    resp = requests.post(WEBAPP_URL, json=payload, timeout=120)
    result = resp.json()
    for r in result.get('results', []):
        icon = "✅" if r['status'] == 'OK' else "⏭️"
        print(f"  {icon} {r['sheet']}: {r['status']} — {r['rows']}행")
    return result

# ========== STEP 2: 종합피벗 갱신 ==========
def step2(mode, target_date):
    print(f"\n=== STEP 2: 종합피벗 갱신 (MODE {mode}) ===")
    
    # 전체 Raw 다운로드
    mh, mr = dl_sheet(GIDS['meta'])
    mi = {h.strip(): i for i, h in enumerate(mh)}
    
    nh, nr = dl_sheet(GIDS['nv_grp'])
    ni = {h.strip(): i for i, h in enumerate(nh)}
    
    # GA 전환/폼유입 계산
    ga_conv, form_starts = {}, {}
    seen = set()
    for ga_key in ['ga1', 'ga2']:
        gh, gr = dl_sheet(GIDS[ga_key])
        for r in gr:
            if len(r) < 5: continue
            dt, ev, sm, camp = str(r[0]).strip()[:10], str(r[1]).strip(), str(r[2]).strip(), str(r[3]).strip()
            users = sf(r[4])
            dk = (dt, ev, sm)
            if dk in seen: continue
            seen.add(dk)
            ch = classify_channel(sm, camp)
            if '최종제출' in ev and ch != 'coupang':
                ga_conv[dt] = ga_conv.get(dt, 0) + users
            if ev == 'form_start' and ch != 'coupang':
                form_starts[dt] = form_starts.get(dt, 0) + users
    
    # 메타 리드폼
    meta_leads = {}
    seen_ml = set()
    for r in mr:
        if len(r) < 10: continue
        if str(r[mi['결과 유형']]).strip() != '잠재 고객(양식)': continue
        dt = str(r[mi['일']]).strip()[:10]
        ad = str(r[mi['광고 이름']]).strip()
        result = sf(r[mi['결과']])
        dk = (dt, ad, result)
        if dk in seen_ml: continue
        seen_ml.add(dk)
        meta_leads[dt] = meta_leads.get(dt, 0) + result
    
    # 당일 데이터 (MODE A)
    td = target_date.strftime('%Y-%m-%d')
    pd = (target_date - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # 메타 당일/전일
    m_day = {'spend':0,'imp':0,'click':0}
    m_prev = {'spend':0,'imp':0,'click':0}
    for r in mr:
        if len(r) < 18: continue
        d = str(r[mi['일']]).strip()[:10]
        if d == td:
            m_day['spend'] += sf(r[mi['지출 금액 (KRW)']])
            m_day['imp'] += sf(r[mi['노출']])
            m_day['click'] += sf(r[mi['링크 클릭']])
        elif d == pd:
            m_prev['spend'] += sf(r[mi['지출 금액 (KRW)']])
            m_prev['imp'] += sf(r[mi['노출']])
            m_prev['click'] += sf(r[mi['링크 클릭']])
    
    # 네이버 당일/전일
    n_day = {'spend':0,'imp':0,'click':0}
    n_prev = {'spend':0,'imp':0,'click':0}
    for r in nr:
        if len(r) < 9: continue
        d = str(r[ni['일별']]).strip()[:10]
        if d == td:
            n_day['spend'] += sf(r[ni['총비용(VAT포함,원)']])
            n_day['imp'] += sf(r[ni['노출수']])
            n_day['click'] += sf(r[ni['클릭수']])
        elif d == pd:
            n_prev['spend'] += sf(r[ni['총비용(VAT포함,원)']])
            n_prev['imp'] += sf(r[ni['노출수']])
            n_prev['click'] += sf(r[ni['클릭수']])
    
    # 당일 합계
    day = {
        'spend': m_day['spend'] + n_day['spend'],
        'imp': m_day['imp'] + n_day['imp'],
        'click': m_day['click'] + n_day['click'],
        'conv': ga_conv.get(td, 0) + meta_leads.get(td, 0),
        'ga_conv': ga_conv.get(td, 0),
        'meta_lead': meta_leads.get(td, 0),
        'fs': form_starts.get(td, 0),
        'meta': m_day, 'naver': n_day,
    }
    prev = {
        'spend': m_prev['spend'] + n_prev['spend'],
        'imp': m_prev['imp'] + n_prev['imp'],
        'click': m_prev['click'] + n_prev['click'],
        'conv': ga_conv.get(pd, 0) + meta_leads.get(pd, 0),
        'fs': form_starts.get(pd, 0),
        'meta': m_prev, 'naver': n_prev,
    }
    
    # 누적 계산 (전체 광고비)
    total_meta_spend = sum(sf(r[mi['지출 금액 (KRW)']]) for r in mr)
    total_meta_imp = sum(sf(r[mi['노출']]) for r in mr)
    total_meta_click = sum(sf(r[mi['링크 클릭']]) for r in mr)
    total_nv_spend = sum(sf(r[ni['총비용(VAT포함,원)']]) for r in nr)
    total_nv_imp = sum(sf(r[ni['노출수']]) for r in nr)
    total_nv_click = sum(sf(r[ni['클릭수']]) for r in nr)
    
    # 당근 (기존 데이터만)
    kh, kr = dl_sheet(GIDS['karrot'])
    total_kr_spend = sum(sf(r[4]) for r in kr if len(r) > 4 and str(r[0]).startswith('20'))
    total_kr_imp = sum(sf(r[5]) for r in kr if len(r) > 5 and str(r[0]).startswith('20'))
    total_kr_click = sum(sf(r[7]) for r in kr if len(r) > 7 and str(r[0]).startswith('20'))
    
    cum_spend = total_meta_spend + total_nv_spend + total_kr_spend
    cum_imp = total_meta_imp + total_nv_imp + total_kr_imp
    cum_click = total_meta_click + total_nv_click + total_kr_click
    cum_conv = sum(ga_conv.values()) + sum(meta_leads.values())
    cum_fs = sum(form_starts.values())
    cum_ctr = cum_click / cum_imp if cum_imp > 0 else 0
    cum_cpc = cum_spend / cum_click if cum_click > 0 else 0
    cum_cpr = cum_spend / cum_conv if cum_conv > 0 else 0
    
    cum = {'spend': cum_spend, 'imp': cum_imp, 'click': cum_click, 'conv': cum_conv,
           'ctr': cum_ctr, 'cpc': cum_cpc, 'cpr': cum_cpr, 'fs': cum_fs}
    
    # 구역1 KPI 업데이트 (브라우저 paste 대신 직접 계산 결과 저장)
    dates = sorted(set(str(r[0]).strip()[:10] for r in mr if str(r[0]).startswith('20')) |
                   set(str(r[ni['일별']]).strip()[:10] for r in nr if str(r[ni['일별']]).startswith('20')))
    period = f"{dates[0].replace('-','')[2:]}-{dates[-1].replace('-','')[2:]}" if dates else ""
    
    kpi_row = [period, fmt(int(cum_spend)), fmt(int(cum_imp)), fmt(int(cum_click)),
               f"{cum_ctr:.2%}", fmt(int(cum_cpc)), fmt(int(cum_cpr)),
               f"{int(cum_conv)}건(GA {int(sum(ga_conv.values()))} + 메타리드 {int(sum(meta_leads.values()))})",
               f"{int(cum_fs)}건"]
    
    print(f"  KPI: 광고비={fmt(int(cum_spend))} 전환={int(cum_conv)} CTR={cum_ctr:.2%}")
    print(f"  당일: 광고비={fmt(int(day['spend']))} 전환={int(day['conv'])} 폼유입={int(day['fs'])}")
    
    # ===== 구역2: 주차별 데이터 (MODE B에서만 전체 갱신, MODE A에서도 KPI는 갱신) =====
    # 주차 계산
    data_start = datetime(2026, 2, 6)
    first_monday = datetime(2026, 2, 9)
    
    def get_week_num(date_str):
        try: d = datetime.strptime(date_str[:10], '%Y-%m-%d')
        except: return -1
        if d < first_monday: return 0
        return (d - first_monday).days // 7 + 1
    
    periods_map = {}
    for w in range(20):
        if w == 0:
            ws_d, we_d = data_start, first_monday - timedelta(days=1)
        else:
            ws_d = first_monday + timedelta(days=(w-1)*7)
            we_d = ws_d + timedelta(days=6)
        wdn = ['월','화','수','목','금','토','일']
        periods_map[w] = f"{ws_d.month}/{ws_d.day}({wdn[ws_d.weekday()]})~{we_d.month}/{we_d.day}({wdn[we_d.weekday()]})"
    
    # 채널별 주차 데이터 집계
    from collections import defaultdict
    meta_wk = defaultdict(lambda: {'spend':0,'imp':0,'click':0})
    nv_wk = defaultdict(lambda: {'spend':0,'imp':0,'click':0})
    kr_wk = defaultdict(lambda: {'spend':0,'imp':0,'click':0})
    
    for r in mr:
        if len(r) < 18: continue
        d_str = str(r[mi['일']]).strip()[:10]
        w = get_week_num(d_str)
        if w < 0: continue
        meta_wk[w]['spend'] += sf(r[mi['지출 금액 (KRW)']])
        meta_wk[w]['imp'] += sf(r[mi['노출']])
        meta_wk[w]['click'] += sf(r[mi['링크 클릭']])
    
    for r in nr:
        if len(r) < 9: continue
        d_str = str(r[ni['일별']]).strip()[:10]
        w = get_week_num(d_str)
        if w < 0: continue
        nv_wk[w]['spend'] += sf(r[ni['총비용(VAT포함,원)']])
        nv_wk[w]['imp'] += sf(r[ni['노출수']])
        nv_wk[w]['click'] += sf(r[ni['클릭수']])
    
    for r in kr:
        if len(r) < 8 or not str(r[0]).startswith('20'): continue
        w = get_week_num(str(r[0]).strip()[:10])
        if w < 0: continue
        kr_wk[w]['spend'] += sf(r[4])
        kr_wk[w]['imp'] += sf(r[5])
        kr_wk[w]['click'] += sf(r[7])
    
    # GA 전환/폼유입 채널별 주차 집계
    ga_conv_wk = defaultdict(lambda: defaultdict(float))
    fs_wk = defaultdict(lambda: defaultdict(float))
    meta_lead_wk = defaultdict(float)
    
    seen2 = set()
    for ga_key in ['ga1', 'ga2']:
        gh, gr = dl_sheet(GIDS[ga_key])
        for r in gr:
            if len(r) < 5: continue
            dt2, ev2, sm2, camp2 = str(r[0]).strip()[:10], str(r[1]).strip(), str(r[2]).strip(), str(r[3]).strip()
            users2 = sf(r[4])
            dk2 = (dt2, ev2, sm2)
            if dk2 in seen2: continue
            seen2.add(dk2)
            ch2 = classify_channel(sm2, camp2)
            w2 = get_week_num(dt2)
            if w2 < 0: continue
            if '최종제출' in ev2 and ch2 != 'coupang':
                ga_conv_wk[w2][ch2] += users2
            if ev2 == 'form_start' and ch2 != 'coupang':
                fs_wk[w2][ch2] += users2
    
    for r in mr:
        if len(r) < 10: continue
        if str(r[mi['결과 유형']]).strip() != '잠재 고객(양식)': continue
        dt2 = str(r[mi['일']]).strip()[:10]
        w2 = get_week_num(dt2)
        if w2 >= 0: meta_lead_wk[w2] += sf(r[mi['결과']])
    
    all_weeks = sorted(set(list(meta_wk.keys()) + list(nv_wk.keys()) + list(kr_wk.keys())))
    
    # ===== 종합피벗 TSV 생성 (전체 교체) =====
    def build_ch_row(channel, w, d_wk, prev_wk, goals=None):
        sp, im, cl = d_wk['spend'], d_wk['imp'], d_wk['click']
        g = goals or {}
        can = w >= 2 and prev_wk
        ctr_v = cl/im if im > 0 else 0
        cpc_v = sp/cl if cl > 0 else 0
        cpm_v = sp/im*1000 if im > 0 else 0
        
        # 전환/폼유입
        if channel == '메타':
            cv = ga_conv_wk[w].get('meta',0) + meta_lead_wk.get(w,0)
            fs_v = fs_wk[w].get('meta',0)
        elif channel == '네이버SA':
            cv = ga_conv_wk[w].get('naver_sa',0)
            fs_v = fs_wk[w].get('naver_sa',0)
        else:
            cv = ga_conv_wk[w].get('karrot',0)
            fs_v = fs_wk[w].get('karrot',0)
        cpr_v = sp/cv if cv > 0 else 0
        
        if channel == '당근' and w >= 7:
            return [channel, f"{w}주차", periods_map.get(w,''), "", "집행 중단 (3/20~)", "—", "", "집행 중단", "—", "", "집행 중단", "—", "", "—", "—", "", "—", "—", "", "—", "—", "", "—", "—", "", "—", "—", "", "—", "—"]
        
        def vchg(cur, prev_val):
            s = fmt(int(cur)) if cur > 0 else "0"
            if can and prev_val > 0:
                c = chg(cur, prev_val)
                if c: return s + c
            return s
        def achv(actual, goal_key):
            gv = g.get(goal_key, 0)
            if not gv: return "—"
            rate = actual / gv * 100
            return f"✅ {rate:.0f}%" if rate >= 100 else f"❌ {rate:.0f}%"
        
        p = prev_wk or {'spend':0,'imp':0,'click':0}
        p_ctr = p['click']/p['imp'] if p.get('imp',0) > 0 else 0
        p_cpc = p['spend']/p['click'] if p.get('click',0) > 0 else 0
        p_cpm = p['spend']/p['imp']*1000 if p.get('imp',0) > 0 else 0
        
        ctr_s = f"{ctr_v:.2%}" + (chg(ctr_v, p_ctr) if can and p_ctr > 0 else "")
        cpc_s = (fmt(int(cpc_v)) + (chg(cpc_v, p_cpc) if can and p_cpc > 0 else "")) if cl > 0 else "—"
        cpm_s = (fmt(int(cpm_v)) + (chg(cpm_v, p_cpm) if can and p_cpm > 0 else "")) if im > 0 else "—"
        cpr_s = fmt(int(cpr_v)) if cv > 0 else "—"
        
        return [channel, f"{w}주차", periods_map.get(w,''),
                fmt(g.get('spend','')) if g.get('spend') else "", vchg(sp, p.get('spend',0)), achv(sp, 'spend'),
                fmt(g.get('imp','')) if g.get('imp') else "", vchg(im, p.get('imp',0)), achv(im, 'imp'),
                fmt(g.get('click','')) if g.get('click') else "", vchg(cl, p.get('click',0)), achv(cl, 'click'),
                "", ctr_s, "—", "", cpc_s, "—", "", cpm_s, "—",
                "", vchg(cv, 0), "—", "", vchg(fs_v, 0), "—", "", cpr_s, "—"]
    
    goals_6 = {"meta":{"spend":550000,"imp":70000,"click":1500}, "naver":{"spend":350000,"imp":12000,"click":30}, "karrot":{"spend":200000,"imp":55000,"click":130}}
    
    pivot_rows = []
    pivot_rows.append(["","","","","마케팅 KPI"]+[""]*(30-5))
    pivot_rows.append(["총 기간","총 광고비","총 노출","총 클릭","평균 CTR","평균 CPC","결과당 비용","전환수","폼 유입"]+[""]*(30-9))
    pivot_rows.append(kpi_row + [""]*(30-len(kpi_row)))
    pivot_rows.append([""]*30)
    pivot_rows.append(["채널별 주간 지표"]+[""]*29)
    pivot_rows.append(["채널","주차","기간","광고비","","","노출수","","","클릭수","","","CTR","","","CPC","","","CPM","","","전환수","","","폼 유입","","","결과당비용","",""])
    pivot_rows.append(["","","","목표","실제","달성률","목표","실제","달성률","목표","실제","달성률","목표","실제","달성률","목표","실제","달성률","목표","실제","달성률","목표","실제","달성률","목표","실제","달성률","목표","실제","달성률"])
    
    for ch_name, wk_data, ch_key, goal_key in [("메타", meta_wk, "meta", "meta"), ("네이버SA", nv_wk, "naver_sa", "naver"), ("당근", kr_wk, "karrot", "karrot")]:
        for w in all_weeks:
            prev_w = wk_data.get(w-1) if w > 0 else None
            g = goals_6.get(goal_key, {}) if w == 6 else {}
            pivot_rows.append(build_ch_row(ch_name, w, wk_data[w], prev_w, g))
        # 합계
        ts = sum(wk_data[w]['spend'] for w in all_weeks)
        ti = sum(wk_data[w]['imp'] for w in all_weeks)
        tc = sum(wk_data[w]['click'] for w in all_weeks)
        tcv = sum(ga_conv_wk[w].get(ch_key,0) for w in all_weeks) + (sum(meta_lead_wk.get(w,0) for w in all_weeks) if ch_name=='메타' else 0)
        tfs = sum(fs_wk[w].get(ch_key,0) for w in all_weeks)
        t_ctr = tc/ti if ti > 0 else 0
        t_cpc = ts/tc if tc > 0 else 0
        t_cpm = ts/ti*1000 if ti > 0 else 0
        t_cpr = ts/tcv if tcv > 0 else 0
        pivot_rows.append([f"{ch_name} 합계","","","",fmt(int(ts)),"","",fmt(int(ti)),"","",fmt(int(tc)),"",
                          "",f"{t_ctr:.2%}","","",fmt(int(t_cpc)) if tc>0 else "—","","",fmt(int(t_cpm)) if ti>0 else "—","",
                          "",fmt(int(tcv)),"","",fmt(int(tfs)),"","",fmt(int(t_cpr)) if tcv>0 else "—",""])
    
    # 전체 합계
    pivot_rows.append(["전체 합계","","","",fmt(int(cum_spend)),"","",fmt(int(cum_imp)),"","",fmt(int(cum_click)),"",
                       "",f"{cum_ctr:.2%}","","",fmt(int(cum_cpc)),"","",fmt(int(cum_spend/cum_imp*1000)) if cum_imp>0 else "—","",
                       "",fmt(int(cum_conv)),"","",fmt(int(cum_fs)),"","",fmt(int(cum_cpr)) if cum_conv>0 else "—",""])
    pivot_rows.append([""]*30)
    pivot_rows.append(["[범례]"]+[""]*29)
    pivot_rows.append(["빨간색(#DC2626) = ▲ 전주대비 상승"]+[""]*29)
    pivot_rows.append(["파란색(#2563EB) = ▼ 전주대비 하락"]+[""]*29)
    pivot_rows.append(["목표 셀(노란색): 직접 입력 → 달성률 자동 계산"]+[""]*29)
    pivot_rows.append(["전환 = GA 최종제출 + 메타 리드폼 (쿠팡 제외, CRM 포함)"]+[""]*29)
    pivot_rows.append(["주차 기준: 월요일 ~ 일요일 (KST 한국시간)"]+[""]*29)
    
    # TSV로 저장 (나중에 브라우저 paste용)
    pivot_tsv = '\n'.join(['\t'.join(str(c) for c in row) for row in pivot_rows])
    pivot_tsv_path = '/tmp/pivot_auto.tsv'
    with open(pivot_tsv_path, 'w', encoding='utf-8') as f:
        f.write(pivot_tsv)
    print(f"  종합피벗 TSV 생성: {len(pivot_rows)}행")
    
    return {'day': day, 'prev': prev, 'cum': cum, 'kpi_row': kpi_row,
            'target_date': td, 'prev_date': pd, 'pivot_tsv': pivot_tsv_path}

# ========== STEP 3: 보고서 생성 ==========
def step3_daily(data, target_date):
    print(f"\n=== STEP 3: 데일리 보고서 생성 ===")
    d = data['day']
    p = data['prev']
    c = data['cum']
    td = target_date.strftime('%Y.%m.%d')
    td_short = target_date.strftime('%m/%d').lstrip('0').replace('/0','/')
    weekdays = ['월','화','수','목','금','토','일']
    wd = weekdays[target_date.weekday()]
    ymd = target_date.strftime('%Y%m%d')[2:]
    
    def pct_chg(cur, prev):
        if prev == 0: return "▲" if cur > 0 else "—"
        ch = (cur - prev) / prev * 100
        if abs(ch) < 0.1: return "—"
        return f"▲ {abs(ch):.1f}%" if ch > 0 else f"▼ {abs(ch):.1f}%"
    
    def css_class(text):
        if '▲' in text: return 'up'
        if '▼' in text: return 'dn'
        return 'mu'
    
    ctr = d['click']/d['imp'] if d['imp'] > 0 else 0
    cpc = d['spend']/d['click'] if d['click'] > 0 else 0
    cpr = d['spend']/d['conv'] if d['conv'] > 0 else 0
    p_ctr = p['click']/p['imp'] if p['imp'] > 0 else 0
    p_cpr = p['spend']/p['conv'] if p['conv'] > 0 else 0
    
    m = d['meta']
    m_ctr = m['click']/m['imp'] if m['imp'] > 0 else 0
    m_cpc = m['spend']/m['click'] if m['click'] > 0 else 0
    m_pct = f"{m['spend']/d['spend']*100:.1f}%" if d['spend'] > 0 else '—'
    n_pct = f"{d['naver']['spend']/d['spend']*100:.1f}%" if d['spend'] > 0 else '—'
    n_ctr = f"{d['naver']['click']/d['naver']['imp']*100:.2f}%" if d['naver']['imp'] > 0 else '—'
    n_cpc = fmt(int(d['naver']['spend']/d['naver']['click'])) if d['naver']['click'] > 0 else '—'
    d_cpr = fmt(int(d['spend']/d['conv'])) if d['conv'] > 0 else '—'
    
    # KPI changes
    chg_spend = pct_chg(d['spend'], p['spend'])
    chg_click = pct_chg(d['click'], p['click'])
    chg_ctr = pct_chg(ctr, p_ctr)
    chg_conv = pct_chg(d['conv'], p['conv'])
    
    filename = f"daily_{ymd}.html"
    filepath = os.path.join(REPORT_DIR, filename)
    
    # HTML 생성 (템플릿)
    html = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>차액가맹금 데일리 리포트 {td}</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:'Noto Sans KR',system-ui,sans-serif;color:#1E293B;background:#F8FAFC;line-height:1.6}}.wrap{{max-width:800px;margin:0 auto;background:#fff}}
.cover{{background:linear-gradient(135deg,#1E293B,#0F172A);color:#fff;padding:48px 40px 36px}}.cover .top{{display:flex;justify-content:space-between;align-items:center;margin-bottom:32px}}.cover .logo{{font-size:8pt;color:#94A3B8;letter-spacing:1px}}.cover .badge{{background:#7B68AE;font-size:7pt;padding:3px 12px;border-radius:10px;letter-spacing:1.5px;text-transform:uppercase}}.cover h1{{font-size:22pt;font-weight:900;line-height:1.2;margin-bottom:4px}}.cover .sub{{font-size:10pt;color:#94A3B8;font-weight:300;margin-bottom:20px}}.cover .divider{{width:36px;height:2px;background:#7B68AE;margin:18px 0 24px}}.kpi-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}}.kpi-card{{background:rgba(255,255,255,.05);border-radius:8px;padding:14px 12px;border-left:3px solid #7B68AE}}.kpi-card .lb{{font-size:6.5pt;color:#94A3B8;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}}.kpi-card .vl{{font-size:18pt;font-weight:700}}.kpi-card .ch{{font-size:7pt;margin-top:3px}}.cover .meta{{font-size:7pt;color:#475569;margin-top:24px;line-height:1.8}}
.up{{color:#DC2626}}.dn{{color:#2563EB}}.mu{{color:#94A3B8}}
.body{{padding:28px 40px 40px}}.sh{{background:#1E293B;border-radius:4px;padding:8px 14px;margin:22px 0 12px;display:flex;align-items:center;gap:8px}}.sh .n{{background:#7B68AE;color:#fff;font-size:7.5pt;font-weight:700;padding:2px 9px;border-radius:3px}}.sh .t{{font-size:10pt;font-weight:700;color:#fff}}.ib{{border-left:3px solid #7B68AE;background:#F8F7FC;padding:12px 16px;margin:8px 0;border-radius:0 4px 4px 0;font-size:8.5pt;line-height:1.7}}
table{{width:100%;border-collapse:collapse;font-size:7.5pt;margin:8px 0}}th{{background:#1E293B;color:#fff;padding:6px 8px;font-weight:600;font-size:7pt;text-align:center}}td{{padding:5px 8px;border-bottom:1px solid #E2E8F0;text-align:center}}tr:nth-child(even) td{{background:#F8FAFC}}.gt td{{background:#1E293B!important;color:#fff;font-weight:700}}td.l{{text-align:left}}
.kpi-body{{display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin:10px 0}}.kb{{background:#1E293B;border-radius:6px;padding:10px 12px}}.kb .lb{{font-size:6.5pt;color:#94A3B8}}.kb .vl{{font-size:14pt;font-weight:700;color:#fff}}.kb .ch{{font-size:7pt}}
.tag{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:6.5pt;font-weight:700}}.p1{{background:#FEE2E2;color:#DC2626}}.p2{{background:#FEF3C7;color:#D97706}}.act{{margin:10px 0 6px}}.act .tt{{font-weight:700;font-size:8.5pt;margin-top:4px}}.act .ds{{font-size:7.5pt;color:#475569;margin-top:2px}}
.warn{{background:#FFFBEB;border-left:3px solid #F59E0B;padding:10px 14px;margin:14px 0;border-radius:0 4px 4px 0;font-size:7.5pt}}.warn b{{color:#D97706;font-size:8pt}}.ft{{margin-top:32px;padding-top:10px;border-top:1px solid #E2E8F0;font-size:6.5pt;color:#94A3B8;text-align:center}}
</style></head><body><div class="wrap">
<div class="cover">
<div class="top"><span class="logo">AI Consilium</span><span class="badge">Daily Report</span></div>
<h1>차액가맹금 집단소송<br>마케팅 데일리 리포트</h1>
<div class="sub">{td} ({wd})</div><div class="divider"></div>
<div class="kpi-row">
<div class="kpi-card"><div class="lb">당일 광고비</div><div class="vl">{fmt(int(d['spend']))}</div><div class="ch {css_class(chg_spend)}">{chg_spend} vs 전일</div></div>
<div class="kpi-card"><div class="lb">당일 클릭</div><div class="vl">{fmt(int(d['click']))}</div><div class="ch {css_class(chg_click)}">{chg_click}</div></div>
<div class="kpi-card"><div class="lb">CTR</div><div class="vl">{ctr:.2%}</div><div class="ch {css_class(chg_ctr)}">{chg_ctr}</div></div>
<div class="kpi-card"><div class="lb">전환수</div><div class="vl">{int(d['conv'])}건</div><div class="ch {css_class(chg_conv)}">{chg_conv}</div></div>
</div>
<div class="meta">분석 날짜: {td} ({wd}) | GA · Meta Ads · Naver SA<br>작성: 에이아이컨실리움 | 수신: 법무법인 더 에이치 황해 | CONFIDENTIAL</div>
</div>
<div class="body">
<div class="sh"><span class="n">D1</span><span class="t">오늘의 핵심 요약</span></div>
<div class="ib"><b>당일 전환 {int(d['conv'])}건</b> (GA {int(d['ga_conv'])} + 메타리드 {int(d['meta_lead'])}). 광고비 {fmt(int(d['spend']))}원. 폼유입 {int(d['fs'])}건(전일 {int(p['fs'])}건).</div>
<div class="sh"><span class="n">D2</span><span class="t">당일 KPI 스냅샷</span></div>
<div class="kpi-body">
<div class="kb"><div class="lb">광고비</div><div class="vl">{fmt(int(d['spend']))}</div><div class="ch {css_class(chg_spend)}">{chg_spend}</div></div>
<div class="kb"><div class="lb">노출</div><div class="vl">{fmt(int(d['imp']))}</div><div class="ch {css_class(pct_chg(d['imp'],p['imp']))}">{pct_chg(d['imp'],p['imp'])}</div></div>
<div class="kb"><div class="lb">클릭</div><div class="vl">{fmt(int(d['click']))}</div><div class="ch {css_class(chg_click)}">{chg_click}</div></div>
<div class="kb"><div class="lb">CTR</div><div class="vl">{ctr:.2%}</div><div class="ch {css_class(chg_ctr)}">{chg_ctr}</div></div>
<div class="kb"><div class="lb">CPC</div><div class="vl">{fmt(int(cpc))}</div><div class="ch {css_class(pct_chg(cpc, p['spend']/p['click'] if p['click']>0 else 0))}">{pct_chg(cpc, p['spend']/p['click'] if p['click']>0 else 0)}</div></div>
<div class="kb"><div class="lb">전환수</div><div class="vl">{int(d['conv'])}건</div><div class="ch {css_class(chg_conv)}">{chg_conv}</div></div>
</div>
<table>
<tr><th>지표</th><th>당일 ({td_short})</th><th>전일</th><th>전일대비</th><th>누적</th></tr>
<tr><td class="l"><b>광고비</b></td><td>{fmt(int(d['spend']))}</td><td>{fmt(int(p['spend']))}</td><td class="{css_class(chg_spend)}">{chg_spend}</td><td>{fmt(int(c['spend']))}</td></tr>
<tr><td class="l"><b>노출</b></td><td>{fmt(int(d['imp']))}</td><td>{fmt(int(p['imp']))}</td><td class="{css_class(pct_chg(d['imp'],p['imp']))}">{pct_chg(d['imp'],p['imp'])}</td><td>{fmt(int(c['imp']))}</td></tr>
<tr><td class="l"><b>클릭</b></td><td>{fmt(int(d['click']))}</td><td>{fmt(int(p['click']))}</td><td class="{css_class(chg_click)}">{chg_click}</td><td>{fmt(int(c['click']))}</td></tr>
<tr><td class="l"><b>CTR</b></td><td>{ctr:.2%}</td><td>{p_ctr:.2%}</td><td class="{css_class(chg_ctr)}">{chg_ctr}</td><td>{c['ctr']:.2%}</td></tr>
<tr><td class="l"><b>CPC</b></td><td>{fmt(int(cpc))}</td><td>{fmt(int(p['spend']/p['click'])) if p['click']>0 else '—'}</td><td class="{css_class(pct_chg(cpc, p['spend']/p['click'] if p['click']>0 else 0))}">{pct_chg(cpc, p['spend']/p['click'] if p['click']>0 else 0)}</td><td>{fmt(int(c['cpc']))}</td></tr>
<tr><td class="l"><b>전환수</b></td><td>{int(d['conv'])}</td><td>{int(p['conv'])}</td><td class="{css_class(chg_conv)}">{chg_conv}</td><td>{int(c['conv'])}</td></tr>
<tr><td class="l"><b>결과당비용</b></td><td>{fmt(int(cpr)) if d['conv']>0 else '—'}</td><td>{fmt(int(p_cpr)) if p['conv']>0 else '—'}</td><td class="{css_class(pct_chg(cpr, p_cpr) if d['conv']>0 and p['conv']>0 else '—')}">{pct_chg(cpr, p_cpr) if d['conv']>0 and p['conv']>0 else '—'}</td><td>{fmt(int(c['cpr']))}</td></tr>
<tr><td class="l"><b>폼유입</b></td><td>{int(d['fs'])}</td><td>{int(p['fs'])}</td><td class="{css_class(pct_chg(d['fs'],p['fs']))}">{pct_chg(d['fs'],p['fs'])}</td><td>{int(c['fs'])}</td></tr>
</table>
<p style="font-size:6.5pt;color:#94A3B8">※ 쿠팡 제외 · CRM 포함 · 당근 3/20~ 중단</p>
<div class="sh"><span class="n">D3</span><span class="t">채널별 당일 성과</span></div>
<table>
<tr><th>채널</th><th>광고비</th><th>비중</th><th>노출</th><th>클릭</th><th>CTR</th><th>CPC</th><th>전환</th><th>결과당비용</th></tr>
<tr><td class="l"><b>메타</b></td><td>{fmt(int(m['spend']))}</td><td>{m_pct}</td><td>{fmt(int(m['imp']))}</td><td>{fmt(int(m['click']))}</td><td>{m_ctr:.2%}</td><td>{fmt(int(m_cpc))}</td><td>{int(d['conv'])}</td><td>{d_cpr}</td></tr>
<tr><td class="l"><b>네이버SA</b></td><td>{fmt(int(d['naver']['spend']))}</td><td>{n_pct}</td><td>{fmt(int(d['naver']['imp']))}</td><td>{fmt(int(d['naver']['click']))}</td><td>{n_ctr}</td><td>{n_cpc}</td><td>0</td><td>—</td></tr>
<tr><td class="l"><b>당근</b></td><td colspan="8" style="color:#94A3B8;text-align:center">광고 중단 (3/20~)</td></tr>
<tr class="gt"><td class="l">합계</td><td>{fmt(int(d['spend']))}</td><td>100%</td><td>{fmt(int(d['imp']))}</td><td>{fmt(int(d['click']))}</td><td>{ctr:.2%}</td><td>{fmt(int(cpc))}</td><td>{int(d['conv'])}</td><td>{fmt(int(cpr)) if d['conv']>0 else '—'}</td></tr>
</table>
<div class="sh"><span class="n">D8</span><span class="t">당일 인사이트 & 액션</span></div>
<div class="ib"><b>메타 전환 {int(d['conv'])}건.</b> 결과당비용 {fmt(int(cpr)) if d['conv']>0 else '—'}원.</div>
<div class="act"><span class="tag p1">P1 즉시</span><div class="tt">전환 0건 소재 점검</div><div class="ds">margin_leadform 등 전환 없는 소재 예산 재배분.</div></div>
<div class="act"><span class="tag p2">P2 내일</span><div class="tt">폼유입 추이 확인</div><div class="ds">당일 {int(d['fs'])}건. 쿠팡 혼입 여부 점검.</div></div>
<div class="warn"><b>데이터 주의사항</b><br>당근 3/20~ 중단 · 쿠팡 제외 · CRM 포함 · 모수 극소</div>
<div class="ft">에이아이컨실리움 | 법무법인 더 에이치 황해 — CONFIDENTIAL | {datetime.now().strftime('%Y-%m-%d')}</div>
</div></div></body></html>"""
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  보고서 생성: {filename}")
    return filename

# ========== GitHub Push ==========
def push_to_github(filename):
    print(f"\n=== GitHub Push ===")
    os.chdir(WORKSPACE)
    subprocess.run(['git', 'add', f'projects/chaek-franchise/reports/{filename}'], capture_output=True)
    subprocess.run(['git', 'commit', '-m', f'보고서 추가: {filename}'], capture_output=True)
    result = subprocess.run(['git', 'push', 'origin', 'main'], capture_output=True, text=True)
    if result.returncode == 0:
        url = f"https://aiconsiliummarketing.github.io/marketing-reports/projects/chaek-franchise/reports/{filename}"
        print(f"  ✅ Push 성공: {url}")
        return url
    else:
        print(f"  ❌ Push 실패: {result.stderr}")
        return None

# ========== 분석 보고서 시트 행 추가 ==========
def add_report_row(target_date, title, url, note):
    print(f"\n=== 분석 보고서 행 추가 ===")
    payload = {
        "sheets": [{
            "name": "분석 보고서",
            "data": [[
                target_date.strftime('%Y-%m-%d'),
                title,
                "메타/네이버SA/GA",
                "에이아이컨실리움",
                url or "",
                note
            ]]
        }]
    }
    resp = requests.post(WEBAPP_URL, json=payload, timeout=30)
    print(f"  결과: {resp.json()}")

# ========== 종합피벗 시트 업데이트 ==========
def update_pivot_sheet(tsv_path):
    if not tsv_path: return
    print(f"\n=== 종합피벗 시트 업데이트 (API) ===")
    
    with open(tsv_path, 'r') as f:
        lines = f.read().strip().split('\n')
    
    data = [line.split('\t') for line in lines]
    for i in range(len(data)):
        while len(data[i]) < 30:
            data[i].append('')
    
    try:
        resp = requests.post(WEBAPP_URL, json={'action': 'updatePivot', 'data': data}, timeout=300)
        result = resp.json()
        if result.get('success'):
            print(f"  ✅ 종합피벗 업데이트 완료: {result.get('rows', 0)}행 + 서식 적용")
        else:
            print(f"  ❌ 실패: {result.get('error', 'unknown')}")
    except Exception as e:
        print(f"  ⚠️ 종합피벗 업데이트 타임아웃 (서버에서 처리 중일 수 있음): {e}")

# ========== MAIN ==========
def run(xlsx_path):
    print(f"\n{'='*50}")
    print(f"차액가맹금 통합 워크플로우 자동 실행")
    print(f"엑셀: {os.path.basename(xlsx_path)}")
    print(f"{'='*50}")
    
    # 모드 판정
    try:
        import pytz
        kst = pytz.timezone('Asia/Seoul')
        now = datetime.now(kst)
    except:
        now = datetime.now() + timedelta(hours=9)
    
    weekday = now.weekday()
    mode = 'B' if weekday == 0 else 'A'
    target_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    # timezone-naive로 변환
    target_date = target_date.replace(tzinfo=None)
    
    wd_names = ['월','화','수','목','금','토','일']
    print(f"\n▶ 오늘: {now.strftime('%Y-%m-%d')} ({wd_names[weekday]}요일) → MODE {'A 데일리' if mode == 'A' else 'B 주간'}")
    print(f"  분석 대상: {target_date.strftime('%Y-%m-%d')}")
    
    # STEP 1
    step1(xlsx_path)
    
    # STEP 2
    data = step2(mode, target_date)
    
    # 종합피벗 시트 업데이트 (TSV paste + formatPivot)
    update_pivot_sheet(data.get('pivot_tsv'))
    
    # STEP 3
    if mode == 'A':
        filename = step3_daily(data, target_date)
        title = f"차액가맹금 마케팅 데일리 리포트 ({target_date.month}/{target_date.day})"
        note = "데일리"
    else:
        # TODO: step3_weekly 구현
        filename = step3_daily(data, target_date)  # 임시
        title = f"차액가맹금 마케팅 주간 보고서"
        note = "주간"
    
    # GitHub Push
    url = push_to_github(filename)
    
    # 분석 보고서 행 추가
    add_report_row(target_date, title, url, note)
    
    print(f"\n{'='*50}")
    print(f"✅ 워크플로우 완료")
    print(f"  보고서: {url}")
    print(f"{'='*50}")
    return url

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 auto_workflow.py <excel_file_path>")
        sys.exit(1)
    run(sys.argv[1])
